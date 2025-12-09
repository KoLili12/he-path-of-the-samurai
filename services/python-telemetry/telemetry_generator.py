#!/usr/bin/env python3
"""
Telemetry Generator - Modern replacement for Pascal legacy service
Generates typed CSV telemetry data and inserts it into PostgreSQL
"""

import os
import sys
import time
import random
import csv
import logging
from datetime import datetime
from typing import List, Dict, Any
import psycopg2
from psycopg2 import sql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)


class TelemetryGenerator:
    """
    Generates telemetry data with proper typing:
    - timestamp (datetime)
    - boolean (true/false)
    - numeric (float)
    - text (string)
    """

    def __init__(self):
        self.csv_out_dir = os.getenv('CSV_OUT_DIR', '/data/csv')
        self.period_sec = int(os.getenv('GEN_PERIOD_SEC', '300'))

        # Database configuration
        self.db_config = {
            'host': os.getenv('PGHOST', 'db'),
            'port': int(os.getenv('PGPORT', '5432')),
            'user': os.getenv('PGUSER', 'monouser'),
            'password': os.getenv('PGPASSWORD', 'monopass'),
            'database': os.getenv('PGDATABASE', 'monolith')
        }

        # Ensure output directory exists
        os.makedirs(self.csv_out_dir, exist_ok=True)

    def generate_telemetry_row(self) -> Dict[str, Any]:
        """
        Generate a single telemetry row with properly typed data

        Returns:
            Dict with keys: recorded_at, voltage, temp, operational, source_file, status
        """
        now = datetime.now()

        return {
            'recorded_at': now,  # TIMESTAMP - exact datetime
            'voltage': round(random.uniform(3.2, 12.6), 2),  # NUMERIC - voltage reading
            'temp': round(random.uniform(-50.0, 80.0), 2),  # NUMERIC - temperature
            'operational': random.choice([True, False]),  # BOOLEAN - system status
            'source_file': f'telemetry_{now.strftime("%Y%m%d_%H%M%S")}.csv',  # TEXT - filename
            'status': random.choice(['OK', 'WARNING', 'ERROR', 'OFFLINE'])  # TEXT - status code
        }

    def write_csv(self, data: List[Dict[str, Any]], filename: str) -> str:
        """
        Write telemetry data to CSV file with proper typing

        Args:
            data: List of telemetry dictionaries
            filename: CSV filename

        Returns:
            Full path to created CSV file
        """
        filepath = os.path.join(self.csv_out_dir, filename)

        try:
            with open(filepath, 'w', newline='') as csvfile:
                fieldnames = ['recorded_at', 'voltage', 'temp', 'operational', 'source_file', 'status']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                writer.writeheader()

                for row in data:
                    # Format data for CSV
                    csv_row = {
                        'recorded_at': row['recorded_at'].isoformat(),  # ISO 8601 format
                        'voltage': f"{row['voltage']:.2f}",  # 2 decimal places
                        'temp': f"{row['temp']:.2f}",  # 2 decimal places
                        'operational': str(row['operational']).upper(),  # TRUE/FALSE
                        'source_file': row['source_file'],  # text
                        'status': row['status']  # text
                    }
                    writer.writerow(csv_row)

            logger.info(f"CSV file created: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Failed to write CSV: {e}")
            raise

    def write_xlsx(self, data: List[Dict[str, Any]], filename: str) -> str:
        """
        Write telemetry data to XLSX file with formatting

        Args:
            data: List of telemetry dictionaries
            filename: XLSX filename

        Returns:
            Full path to created XLSX file
        """
        filepath = os.path.join(self.csv_out_dir, filename.replace('.csv', '.xlsx'))

        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Telemetry Data"

            # Define headers
            headers = ['Recorded At', 'Voltage (V)', 'Temperature (°C)', 'Operational', 'Source File', 'Status']

            # Style for header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Write and style headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write data rows
            for row_num, row_data in enumerate(data, 2):
                ws.cell(row=row_num, column=1, value=row_data['recorded_at'].strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=2, value=row_data['voltage'])
                ws.cell(row=row_num, column=3, value=row_data['temp'])
                ws.cell(row=row_num, column=4, value="YES" if row_data['operational'] else "NO")
                ws.cell(row=row_num, column=5, value=row_data['source_file'])
                ws.cell(row=row_num, column=6, value=row_data['status'])

                # Color code status column
                status_cell = ws.cell(row=row_num, column=6)
                if row_data['status'] == 'OK':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif row_data['status'] == 'WARNING':
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif row_data['status'] == 'ERROR':
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(filepath)
            logger.info(f"XLSX file created: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Failed to write XLSX: {e}")
            raise

    def insert_to_database(self, data: List[Dict[str, Any]]) -> None:
        """
        Insert telemetry data into PostgreSQL using proper types

        Args:
            data: List of telemetry dictionaries
        """
        try:
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor()

            # Create table if not exists with proper column types
            create_table_query = """
            CREATE TABLE IF NOT EXISTS telemetry_legacy (
                id SERIAL PRIMARY KEY,
                recorded_at TIMESTAMPTZ NOT NULL,
                voltage NUMERIC(5,2) NOT NULL,
                temp NUMERIC(5,2) NOT NULL,
                operational BOOLEAN NOT NULL,
                source_file TEXT NOT NULL,
                status TEXT NOT NULL,
                created_at TIMESTAMPTZ DEFAULT NOW()
            );
            """
            cursor.execute(create_table_query)

            # Insert data using parameterized query (SQL injection safe)
            insert_query = """
            INSERT INTO telemetry_legacy
                (recorded_at, voltage, temp, operational, source_file, status)
            VALUES
                (%s, %s, %s, %s, %s, %s)
            """

            for row in data:
                cursor.execute(insert_query, (
                    row['recorded_at'],
                    row['voltage'],
                    row['temp'],
                    row['operational'],
                    row['source_file'],
                    row['status']
                ))

            conn.commit()
            logger.info(f"Inserted {len(data)} rows into database")

            cursor.close()
            conn.close()

        except psycopg2.Error as e:
            logger.error(f"Database error: {e}")
            raise
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            raise

    def generate_and_save(self) -> None:
        """
        Main generation cycle:
        1. Generate telemetry data
        2. Write to CSV and XLSX
        3. Insert into database
        """
        try:
            # Generate multiple rows for batch insert
            num_rows = random.randint(1, 10)
            data = [self.generate_telemetry_row() for _ in range(num_rows)]

            # Use first row's timestamp for filename
            timestamp = data[0]['recorded_at'].strftime('%Y%m%d_%H%M%S')
            filename = f'telemetry_{timestamp}.csv'

            # Update all rows with same filename
            for row in data:
                row['source_file'] = filename

            # Write CSV
            csv_path = self.write_csv(data, filename)

            # Write XLSX (formatted Excel file)
            xlsx_path = self.write_xlsx(data, filename)

            # Insert to database
            self.insert_to_database(data)

            logger.info(f"✓ Successfully generated {num_rows} telemetry records (CSV + XLSX)")

        except Exception as e:
            logger.error(f"✗ Generation cycle failed: {e}")

    def run(self) -> None:
        """
        Main run loop - generates telemetry on schedule
        """
        logger.info(f"Starting Telemetry Generator")
        logger.info(f"CSV output directory: {self.csv_out_dir}")
        logger.info(f"Generation period: {self.period_sec} seconds")
        logger.info(f"Database: {self.db_config['host']}:{self.db_config['port']}/{self.db_config['database']}")

        while True:
            try:
                self.generate_and_save()
                time.sleep(self.period_sec)

            except KeyboardInterrupt:
                logger.info("Shutting down gracefully...")
                break
            except Exception as e:
                logger.error(f"Error in main loop: {e}")
                time.sleep(self.period_sec)


def main():
    """Entry point"""
    generator = TelemetryGenerator()
    generator.run()


if __name__ == '__main__':
    main()
